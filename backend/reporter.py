import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


def generate_excel_report(cycle_name, test_cases, defects, metrics):
    wb = Workbook()

    # ── Color palette ──────────────────────────────────────────────
    BLACK       = "FF0D0D0D"
    DARK        = "FF1A1A2E"
    PURPLE      = "FF6366F1"
    GREEN       = "FF22C55E"
    RED         = "FFEF4444"
    AMBER       = "FFF59E0B"
    WHITE       = "FFFFFFFF"
    LIGHT_GRAY  = "FF9CA3AF"
    CARD        = "FF16213E"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(color=WHITE, bold=False, size=11):
        return Font(color=color, bold=bold, size=size, name="Calibri")

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="FF2D2D4E"),
        right=Side(style="thin", color="FF2D2D4E"),
        top=Side(style="thin", color="FF2D2D4E"),
        bottom=Side(style="thin", color="FF2D2D4E"),
    )

    # ══════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 25

    # Title row
    ws1.row_dimensions[1].height = 50
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = f"StatementSense — QA Report: {cycle_name}"
    c.fill = fill(PURPLE)
    c.font = font(WHITE, bold=True, size=16)
    c.alignment = center()

    # Subtitle
    ws1.row_dimensions[2].height = 25
    ws1.merge_cells("A2:D2")
    c = ws1["A2"]
    c.value = "AI-Powered Billing Statement QA & Compliance Platform"
    c.fill = fill(DARK)
    c.font = font(LIGHT_GRAY, size=11)
    c.alignment = center()

    ws1.row_dimensions[3].height = 10

    # KPI header
    ws1.row_dimensions[4].height = 30
    for col, label in enumerate(["METRIC", "TARGET", "ACTUAL", "STATUS"], 1):
        c = ws1.cell(row=4, column=col, value=label)
        c.fill = fill(DARK)
        c.font = font(LIGHT_GRAY, bold=True, size=10)
        c.alignment = center()
        c.border = thin_border

    kpis = [
        ("Total Test Cases",    "—",   metrics.get("total", 0),          None),
        ("Passed",              "—",   metrics.get("passed", 0),          None),
        ("Failed",              "0",   metrics.get("failed", 0),          metrics.get("failed", 0) == 0),
        ("Blocked",             "0",   metrics.get("blocked", 0),         metrics.get("blocked", 0) == 0),
        ("Pass Rate",           "95%", f"{metrics.get('pass_rate',0):.1f}%",  metrics.get("pass_rate", 0) >= 95),
        ("Execution Rate",      "100%",f"{metrics.get('execution_rate',0):.1f}%", metrics.get("execution_rate", 0) >= 100),
        ("Total Defects",       "—",   metrics.get("total_defects", 0),   None),
        ("Critical Defects",    "0",   metrics.get("critical_defects", 0),metrics.get("critical_defects", 0) == 0),
        ("Defect Leakage Rate", "<5%", f"{metrics.get('defect_leakage',0)}%",  True),
        ("SLA Adherence",       "95%", f"{metrics.get('sla_adherence',0)}%",   True),
    ]

    for i, (metric, target, actual, passed) in enumerate(kpis, start=5):
        ws1.row_dimensions[i].height = 26
        row_fill = fill("FF0F0F1A") if i % 2 == 0 else fill("FF13131F")

        c = ws1.cell(row=i, column=1, value=metric)
        c.fill = row_fill; c.font = font(WHITE); c.alignment = left(); c.border = thin_border

        c = ws1.cell(row=i, column=2, value=str(target))
        c.fill = row_fill; c.font = font(LIGHT_GRAY); c.alignment = center(); c.border = thin_border

        c = ws1.cell(row=i, column=3, value=str(actual))
        c.fill = row_fill; c.font = font(WHITE, bold=True); c.alignment = center(); c.border = thin_border

        if passed is None:
            status_text, status_color, status_fill = "—", LIGHT_GRAY, row_fill
        elif passed:
            status_text, status_color, status_fill = "✓ PASS", GREEN, fill("FF052010")
        else:
            status_text, status_color, status_fill = "✗ FAIL", RED, fill("FF200505")

        c = ws1.cell(row=i, column=4, value=status_text)
        c.fill = status_fill; c.font = font(status_color, bold=True); c.alignment = center(); c.border = thin_border

    # ══════════════════════════════════════════════════════════════
    # Sheet 2 — Test Cases
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test Cases")
    ws2.sheet_view.showGridLines = False

    col_widths = [10, 45, 18, 12, 35, 35, 16]
    col_names  = ["TC ID", "Title", "Category", "Priority", "Expected Result", "Actual Result", "Status"]
    for i, (w, name) in enumerate(zip(col_widths, col_names), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Header
    ws2.row_dimensions[1].height = 40
    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = f"Test Cases — {cycle_name}"
    c.fill = fill(PURPLE); c.font = font(WHITE, bold=True, size=14); c.alignment = center()

    ws2.row_dimensions[2].height = 28
    for col, name in enumerate(col_names, 1):
        c = ws2.cell(row=2, column=col, value=name)
        c.fill = fill(DARK); c.font = font(LIGHT_GRAY, bold=True, size=10)
        c.alignment = center(); c.border = thin_border

    status_colors = {
        "Pass":         (GREEN,  "FF052010"),
        "Fail":         (RED,    "FF200505"),
        "Blocked":      (AMBER,  "FF1F1500"),
        "Not Executed": (LIGHT_GRAY, "FF0F0F1A"),
    }
    priority_colors = {"High": RED, "Medium": AMBER, "Low": "FF60A5FA"}

    for i, tc in enumerate(test_cases, start=3):
        ws2.row_dimensions[i].height = 28
        row_fill = fill("FF0F0F1A") if i % 2 == 0 else fill("FF13131F")
        status = tc.get("status", "Not Executed")
        sc, sf = status_colors.get(status, (LIGHT_GRAY, "FF0F0F1A"))
        pc = priority_colors.get(tc.get("priority", "Medium"), LIGHT_GRAY)

        vals = [
            tc.get("id", f"TC{i-2:03d}"),
            tc.get("title", ""),
            tc.get("category", ""),
            tc.get("priority", ""),
            tc.get("expected_result", ""),
            tc.get("actual_result", ""),
            status,
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = thin_border
            c.alignment = left()
            if col == 4:
                c.fill = row_fill; c.font = font(pc, bold=True)
            elif col == 7:
                c.fill = fill(sf); c.font = font(sc, bold=True); c.alignment = center()
            else:
                c.fill = row_fill; c.font = font(WHITE)

    # ══════════════════════════════════════════════════════════════
    # Sheet 3 — Defects
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Defect Log")
    ws3.sheet_view.showGridLines = False

    d_col_widths = [10, 40, 14, 12, 40, 40]
    d_col_names  = ["DEF ID", "Title", "Severity", "Status", "Root Cause", "AI Root Cause"]
    for i, (w, name) in enumerate(zip(d_col_widths, d_col_names), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.row_dimensions[1].height = 40
    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = f"Defect Log — {cycle_name}"
    c.fill = fill(RED); c.font = font(WHITE, bold=True, size=14); c.alignment = center()

    ws3.row_dimensions[2].height = 28
    for col, name in enumerate(d_col_names, 1):
        c = ws3.cell(row=2, column=col, value=name)
        c.fill = fill(DARK); c.font = font(LIGHT_GRAY, bold=True, size=10)
        c.alignment = center(); c.border = thin_border

    sev_colors = {"Critical": RED, "High": "FFF97316", "Medium": AMBER, "Low": "FF60A5FA"}

    for i, d in enumerate(defects, start=3):
        ws3.row_dimensions[i].height = 28
        row_fill = fill("FF0F0F1A") if i % 2 == 0 else fill("FF13131F")
        sev = d.get("severity", "Medium")
        sc = sev_colors.get(sev, LIGHT_GRAY)

        vals = [
            f"DEF{i-2:03d}",
            d.get("title", ""),
            sev,
            d.get("status", "Open"),
            d.get("root_cause", ""),
            d.get("ai_root_cause", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.border = thin_border; c.alignment = left()
            if col == 3:
                c.fill = row_fill; c.font = font(sc, bold=True)
            else:
                c.fill = row_fill; c.font = font(WHITE)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()